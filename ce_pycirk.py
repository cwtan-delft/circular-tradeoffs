# -*- coding: utf-8 -*-
"""
@author: tanch
"""

import pandas as pd
import numpy as np
import time
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pycirk
from datetime import datetime

import matplotlib.pyplot as plt
# import altair as alt
import seaborn as sns
#%% 
'''
Main function to create analysis table
'''
### create DataFrame of pycirk analyse outputs for use with scenario.xlsx file ###
def create_analyse(category_consumption,region_consumption='All',
                    category_production='All',region_production='All',
                    disaggregate=[]):
    
    ## Base testing directory ##
    folder_data_main = r".\ref"

    ## reading Pycirk "Analyse" sheet ##
    names_categories = pd.read_excel(folder_data_main+
                                     '\scenarios_template.xlsx',
                                     "names_categories",
                                     header=0)
    regions = pd.read_excel(folder_data_main+
                            '\scenarios_template.xlsx','Regions',
                            header=0)
    concordance_cat_mat = pd.read_excel(folder_data_main+
                                        "\Concordance Cat Mat.xlsx",
                                        header = 0, index_col=(0))
    
    ## conduct sanity checks ##
    disagg_accepted = (1,2,3,4)
    cat_sanity = category_sanity(names_categories, category_consumption, 
                                 category_production)
    reg_sanity = region_sanity(regions, region_consumption, region_production)
    
    for i in disaggregate:
        if i not in disagg_accepted:
            print(f"{type(i)}, {i}")
            raise ValueError(
                "disaggregate argument only accepts list containing integers 1-4")
        elif not cat_sanity:
            raise Exception(
                "wrong values for category_production and/or category_consumption")
            return
        elif not reg_sanity:
            raise Exception(
                "wrong values for region_production and/or region_consumption")
            return
        
    
    ## set up category and concordance variables
    category = names_categories[
        names_categories.eq(category_consumption).any(1)].Category.values[0]
    concordance = concordance_cat_mat.loc[category]
    fd_list = ['F_GOVE', 'F_HOUS', 'F_NPSH', 'I_CHIN','I_CHVA','I_EXP', 'I_GFCF']
    
    ## retrieve final demand matrix value
    final = concordance_cat_mat.loc[category].Final
    
    ## create disaggregation list for product categories
    # add 200 product categories
    disagg_products = names_categories[
        names_categories.Category == 'Products'].Abbreviation.values
    # re-add 'All'
    disagg_products = np.insert(disagg_products, 0, 'All')
    # add 7 final demand categories
    if final != "-":
        disagg_products = np.append(disagg_products, fd_list)
    
    ## create disaggregation list for regions
    disagg_regions = regions[0:49].Abbreviation.values 
    
    ## perform disaggragation ##
    if category_consumption == "All" and 1 in disaggregate:
        o_p_list = disagg_products
    else: 
        o_p_list = [category_consumption]
    if category_production == "All" and 3 in disaggregate:
        d_p_list = disagg_products
    else:
        d_p_list = [category_production]
    if region_consumption == "All" and 2 in disaggregate:
        o_r_list = disagg_regions
    else:
        o_r_list = [region_consumption]
    if region_production == "All" and 4 in disaggregate:
        d_r_list = disagg_regions 
    else:
        d_r_list = [region_production]
    
    
    ## create array for writing into analysis sheet
    row = []
    
    for cat_c in o_p_list:
        for reg_c in o_r_list:
            for cat_p in d_p_list:
                for reg_p in d_r_list:

                    ## assign variables corresponding to cell values for rows in analyse table ##
                    o_p = names_categories[names_categories.eq(cat_c).any(1)
                        ].Abbreviation.values[0]
                    o_r = regions[
                        regions.eq(reg_c).any(1)
                        ].Abbreviation.values[0]
                    d_p = names_categories[names_categories.eq(cat_p).any(1)
                        ].Abbreviation.values[0]
                    d_r = regions[regions.eq(reg_p).any(1)
                        ].Abbreviation.values[0]
                    desc = names_categories[names_categories.eq(cat_c).any(1)
                        ].Name.values[0]
                    
                                        
                    ## conduct concordance matching ##
                    ## create array for rows in analyse table ##
                    if cat_p == 'All':
                        for matrix in concordance:
                            if matrix != '-':
                                row.append([matrix,
                                    o_p,o_r,d_p,d_r,f"{desc}: {d_r}/{d_p}"])
                    elif cat_p in fd_list:
                        matrix = concordance['Final']
                        row.append([matrix, 
                                    o_p,o_r,d_p,d_r,f"{desc}: {d_r}/{d_p}"])
                    else:
                        matrix = concordance['Intermediate']
                        row.append([matrix, 
                                    o_p,o_r,d_p,d_r,f"{desc}: {d_r}/{d_p}"])
                            
    return pd.DataFrame(row)

#%%
'''
Sanity check functions
'''                        
### sanity check ###
def category_sanity(names_categories, category_consumption,category_production):
    ## category consumption cannot contain final demand categories
    fd_names = names_categories[names_categories['Category']=='Final demand']
    sanity = None
    
    if fd_names['Category'].isin([category_consumption]).any():
        raise ValueError(
            "category_consumption cannot contain final demand categories, please check")
        sanity = False
    else:
        sanity = True   
    
    ## category production cannot contain W, E, R, M categories
    prod_excl_names = names_categories[
        names_categories['Category'].str.contains(r'Characterisation|extensions')]
    if prod_excl_names['Category'].isin([category_production]).any():
        raise ValueError(
            "category_production cannot contain characterisation or extension categories, please check")
        sanity = False
    else:
        sanity = True
    
    return sanity

def region_sanity(regions, region_consumption,region_production):
    ## check region part of 49 EXIOBase regions
    if not regions.isin([region_consumption]).any().any():
        raise ValueError(
            f"region_consumption: '{region_consumption}' does not contain valid region, please check")
        sanity = False
    else:
        sanity = True
    if not regions.isin([region_production]).any().any():
        raise ValueError(f"region_production: '{region_production}' does not contain valid region, please check")
        sanity = False
    else:
        sanity = True
        
    return sanity

#%%
'''
Write analyse table to scenario.xlsx
'''

def update_analyse(scenario_file, analyse_df):
    ## update analyse sheet in provided scenario.xlsx spreadsheet file
    df_col = len(analyse_df.columns)
    ## sanity check for analyse_df shape
    if df_col != 6:
        raise Exception(f"6 columns expected in Dataframe, found {df_col}")
    else:
        print("Writing analyse to scenarios.xlsx...")
       
        wb = load_workbook(scenario_file)
        ws = wb['analyse']
        ## clear previous analyse rows
        ws.delete_rows(5,ws.max_row+1)
        
        for row in dataframe_to_rows(analyse_df, False, False):
            ws.append(row)
        
        wb.save(scenario_file)
        wb.close() # close openpyxl workbook
        print("writing complete")

#%%
'''
Run pycirk function
'''
def run_pycirk(folder, save=False):
    #### Initialise Pycirk with Launch #####

    timestart = time.time()
    print("starting Pycirk Launch")
    launch = pycirk.Launch(0,True,folder,0,None,False)
    timeend = time.time()
    print("Launch completed in {}s".format(timeend-timestart))

    ## Save analysis results as excel ##
    timestart = time.time()
    print("starting Pycirk results...")
    if save:
        results = launch.save_results()
    else:
        results = launch.all_results()

    timeend = time.time()
    print("Pycirk results - results completed in {}s".format(timeend-timestart))
    
    return results
#%%
'''
Save pycirk data into excel
'''
def save_results_to_excel(data, split_cols=[]):
    now = datetime.now().strftime('%y%m%d_%H.%M')
    scenario_file = f".\\output\\analyse_results_{now}.xlsx"
    
    with pd.ExcelWriter(scenario_file, 'openpyxl', mode='w') as writer:
        if type(data) == dict:
            for k,v in data.items():
                v.to_excel(writer,sheet_name = f'{k}',
                           header=False,merge_cells=False)
                for col in split_cols:
                    for cat in v.index.unique(level=col):
                        df = v[v.index.isin([cat],level=col)]
                        df.to_excel(writer,sheet_name = f'{k}_{cat[:6]}',header=False,merge_cells=False)
        elif type(data) == pd.core.frame.DataFrame:
            data.to_excel(writer,sheet_name = 'results',
                          header=False,merge_cells=False)
            for col in split_cols:
                for cat in data.index.unique(level=col):
                    df = data[data.index.isin([cat],level=col)]
                    df.to_excel(writer,sheet_name = f'result_{cat[:6]}',
                                header=False,merge_cells=False)

#%% melt function
def melt_results(df):
    melted = df.melt(ignore_index=False).reset_index(level=['g_category','i_category'])
    melted.rename(columns= {'variable':'scenario', 'value':'change from baseline'}, 
                  inplace = True)
    melted[['scenario_type','level']] = melted['scenario'].str.split(': ',1, expand = True)
    
    return melted

#%% percentage change calc
def pct_change(df):
    percent_df = df.copy()
    percent_df.iloc[:,1:] = percent_df.iloc[:,1:].subtract(percent_df.baseline,
                                                           axis=0)
    percent_df.iloc[:,1:] = percent_df.iloc[:,1:].div(percent_df.baseline/100, 
                                                      axis=0)
    percent_df.drop(columns=["baseline"], inplace=True)
    
    return percent_df

#%% raw change calc
def raw_change(df):
    raw_df = df.copy()
    raw_df.iloc[:,1:] = raw_df.iloc[:,1:].subtract(raw_df.baseline, axis=0)
    raw_df.drop(columns=["baseline"], inplace=True)
    
    return raw_df

#%%
def drop_all(df, category):
    drop_df = df.copy()
    drop_df = drop_df[drop_df[category] != 'All']
    
    return drop_df

#%% update result df headers
def update_headers(result_df, header_list, baseline=True):
    headers = header_list.copy()
    if baseline:
        headers.insert(0, 'baseline')
    #update headers
    update = result_df.set_axis(headers,axis='columns',inplace=True)        
     
    return update
#%% change domestic extraction to material extraction in MultiIndex
def update_dom_extr(result_df):
    result_df.index = result_df.index.set_levels(
        result_df.index.levels[1].str.replace(
            'Domestic Extraction', 'Material Extraction'), level=1)

    return result_df
#%% sns striplot

def stripplot(df, percent = True, savestring = None):
    if percent:
        processed_df = pct_change(df)
        titletext = ' - percent'
        savestring = savestring+'_percent'
    else:
        processed_df = raw_change(df) 
        titletext = ' - absolute values'
        savestring1 = savestring+'_raw'
    melted = melt_results(processed_df)
    
    for impact in melted['i_category'].unique():
        data = melted.query('i_category == @impact')
        data = drop_all(data, 'g_category')
        # return data
        fig, ax = plt.subplots(figsize=(15,10))
        fig = sns.stripplot(x = 'scenario', y = 'change from baseline', 
                            linewidth = 0.5, jitter = 0.25, palette='Set2', 
                            data = data, ax = ax)
        
        fig.set_xticklabels(fig.get_xticklabels(),rotation = 45,
                                  horizontalalignment = 'right')
        
        ticknames = data.scenario.unique()
        for idx, val in data.iterrows():
            y = val['change from baseline']
            if abs(y) > 15:
                x = ticknames.tolist().index(val['scenario'])
                fig.annotate(val['g_category'], xy=(x+0.1,y), ha = 'left', 
                             va = 'center')
        
        sns.set_context("paper")
        sns.set_theme(font_scale=0.8, style=('ticks'))
        fig.set_title(impact+titletext)
            
        if savestring:
            filename = f'{savestring1}_stripplot.png'
            fig.savefig(filename)
            print(f'saving {filename}')
        
        plt.close()

#%% ana grouped bar (category impacts)

def category_bar(df,percent= True, savestring = None):
    ## process data
    impacttypes= df.index.unique(level = 'i_category')
    scenarios = df.columns
    
    for impact in impacttypes:
        filter_impact = df.query('i_category == @impact')
        unit = filter_impact.index.unique(level = 'unit')[0]
        
        if percent:
            processed_df = pct_change(filter_impact)
            titletext = f'{impact} - percent'
            axistext = 'Percent change of footprint from baseline (%)'
            savestring1 = savestring+f'_{impact}_percent'
        else:
            processed_df = raw_change(filter_impact) 
            titletext = f'{impact} - absolute values'
            axistext = f'Absolute change of footprint from baseline ({unit})'
            savestring1 = savestring+f'_{impact}_raw'
        ## check for positive values
        melted = melt_results(processed_df)
        
        selection = pd.DataFrame()
        for col in scenarios:
            len_positive = len(melted.query(
                'scenario == @col and `change from baseline` > 0'))
            bottom = melted.query('scenario == @col').sort_values(
                'change from baseline', ascending = True).head(5)
            if len_positive >0:
                top = melted.query('scenario == @col').sort_values(
                    'change from baseline', ascending = True, 
                    na_position = 'first'
                    ).tail(min(5, len_positive,))
                bottom = pd.concat([bottom,top])
            selection = pd.concat([selection,bottom])
            

        # for impact in selection['i_category'].unique():
            # data = selection.query('i_category == @impact')
        data = drop_all(selection, 'g_category')
        # fig, ax = plt.subplots()
        cat_plot = sns.catplot(y = 'g_category', x = 'change from baseline', 
                          col = 'scenario', kind='bar', 
                          height=3, aspect=1, col_wrap=3,
                          row_order= scenarios,
                          sharey = False, sharex = False,  palette='coolwarm',
                          data = data)
        sns.set_context("paper")
        sns.set_theme(font_scale=0.8, style=('ticks'))
        
        # annotate
        for ax in cat_plot.axes:
            for c in ax.containers:
                ax.bar_label(c, label_type='edge', fmt='%0.2f')
                # pad the spacing between the number and the edge of the figure
                ax.margins(x=0.4)
                
        cat_plot.fig.subplots_adjust(top=.95)
        cat_plot.fig.suptitle(titletext)
        cat_plot.set(ylabel='product category',xlabel=axistext)
        plt.tight_layout()
        
        if savestring:
            # fig = cat_plot.get_figure()
            fig = cat_plot
            filename = f'{savestring1}_catbar.png'
            fig.savefig(filename)
            print(f'saving {filename}')
            
        plt.close()

#%% sns grouped bar (impact scenarios)
def impact_scenarios_bar(df,savestring = None):
    percent_df = pct_change(df)
    scenario_list = percent_df.columns
    concordance_cat_mat = pd.read_excel(r".\ref\Concordance Cat Mat.xlsx",
                                        header = 0, index_col=(0))
    
    ## check for positive values
    melted = melt_results(percent_df)
    selection = melted[melted["g_category"] == "All"]
    selection = selection[
        selection.index.isin(concordance_cat_mat.Total, 
                             level='matrix')].sort_values(
            "change from baseline", ascending= False)
    
    
    # return selectionn
    cat_plot = sns.catplot(y = 'i_category', x = 'change from baseline', 
                           col = 'scenario', kind='bar', 
                      height=3, aspect=1, col_wrap=3, 
                      col_order= scenario_list,
                      sharey = False, sharex = False, palette='YlGnBu_r',
                      data = selection)
    sns.set_context("paper")
    sns.set_theme(font_scale=0.8, style=('ticks'))
    cat_plot.set(xlabel='Percentage change from baseline (%)',
                 ylabel= 'Impact Category')
    
    # annotate
    for ax in cat_plot.axes:
        for c in ax.containers:
            ax.bar_label(c, label_type='edge', fmt='%0.2f')
            # pad the spacing between the number and the edge of the figure
            ax.margins(x=1)
            
    cat_plot.fig.subplots_adjust(top=.9)
    
    plt.tight_layout()
    # cat_plot.fig.suptitle(impact)
    if savestring:
        # fig = cat_plot.get_figure()
        fig = cat_plot
        filename = f'{savestring}_imapctscenbar.png'
        fig.savefig(filename)
        print(f'saving {filename}')
        
    plt.close()

#%% sns grouped bar (scenario impacts)
def scenario_impacts_bar(df, percent= True, savestring = None):
    if percent:
        processed_df = pct_change(df)
        titletext = ' - percent'
        axistext = '(%)'
        savestring1 = savestring+'_percent'
    else:
        processed_df = raw_change(df) 
        titletext = ' - absolute values'
        savestring1 = savestring+'_raw'

    impact_list = processed_df.index.unique(level='i_category')
    concordance_cat_mat = pd.read_excel(r".\ref\Concordance Cat Mat.xlsx",
                                        header = 0, index_col=(0))
    
    ## check for positive values
    melted = melt_results(processed_df)
    selection = melted[melted["g_category"] == "All"]
    selection = selection[
        selection.index.isin(
            concordance_cat_mat.Total, level='matrix')].sort_values(
            "change from baseline", ascending= False)
    
    
    # return selectionn
    cat_plot = sns.catplot(y = 'scenario', x = 'change from baseline',
                        hue = 'level', 
                        col = 'i_category', kind='bar', 
                        height=5, aspect=0.8, col_wrap=3, 
                        col_order= impact_list,
                        sharey = False, sharex = False, 
                        palette='YlGnBu_r',
                        data = selection)
    sns.set_context("paper")
    sns.set_theme(font_scale=0.8, style=('ticks'))
    
    # annotate
    for ax in cat_plot.axes:
        for c in ax.containers:
            ax.bar_label(c, label_type='edge', fmt='%0.2f')
            # pad the spacing between the number and the edge of the figure
            ax.margins(x=0.5)
            
    cat_plot.fig.subplots_adjust(top=.9)
    cat_plot.fig.suptitle("Impacts"+titletext)
    plt.tight_layout()
    
    if savestring:
        # fig = cat_plot.get_figure()
        fig = cat_plot
        filename = f'{savestring1}_scenimpactbar.png'
        fig.savefig(filename)
        print(f'saving {filename}')
    
    plt.close()
